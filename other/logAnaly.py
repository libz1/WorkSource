#!/usr/bin/python
# -*- coding: utf-8 -*-

import sys
from openpyxl import Workbook

# https://blog.csdn.net/lishuhuakai/article/details/53946091
def dict2list(dic:dict):
    ''' 将字典转化为列表 '''
    keys = dic.keys()
    vals = dic.values()
    lst = [(key, val) for key, val in zip(keys, vals)]
    return lst

def wsWrite(ws,i,line,index):
    ''' 写入excel '''
    dateTime = line[0:23]  #获取日期信息
    ws.cell(row=i, column=1).value = dateTime
    type = line[25:30]  #获取类型信息
    if "2" in type or "3" in type :
       type = "├─"+type
    else:
       if "4" in type:
           type = "└─"+type
       else:
           type = type
    ws.cell(row=i, column=2).value = type
    comID = line[34:36]  #获取端口信息
    ws.cell(row=i, column=3).value = comID 
    ws.cell(row=i, column=4).value = line[36:len(line)].strip() #获取数据信息
    ws.cell(row=i, column=5).value = line
    ws.cell(row=i, column=6).value = index

wb = Workbook()
ws = wb.active

ws.column_dimensions['A'].width = 25
ws.column_dimensions['D'].width = 60

f = open("DEBUG_log.log","r")
line = f.readline()
ws['A1'] = "时间"
ws['B1'] = "操作"
ws['C1'] = "端口"
ws['D1'] = "内容"

dict = {'ZZZZ': 'ZZZZ'}

i = 2
while line:
    if "端口" in line:
       wsWrite(ws,i,line,0)
       
       dateTime = line[0:23]  #获取日期信息
       comID = line[34:36]  #获取端口信息
       type = line[25:30]  #获取类型信息
       key = comID+"^"+dateTime+type
       dict[key] = line
       
       i = i+1
    line = f.readline()

f.close
file = '所有交互过程.xlsx'
wb.save(file)
print("完整日志输出至"+file)

# 数据保存到dict中
# dict转为list ， 对list按照第一个元素进行排序
list = sorted(dict2list(dict), key=lambda x:x[0], reverse=False)

# 根据1出现的次序，定位问题数据
optBegin_idx0 = -1
data = [[-1, -1]]
for i in range(len(list)):
    key = list[i][0]
    value = list[i][1]
    #print(str(i)+"^"+key)
    if "^" in key: 
        type = value[25:30]
        #print(value)
        
        if "1" in type:
            optBegin_idx = i
            #print("optBegin_idx="+str(optBegin_idx))
            if optBegin_idx0 != -1:
                if (optBegin_idx - optBegin_idx0) != 4:
                    #print("find err "+str(optBegin_idx))
                    for j in range(optBegin_idx0,optBegin_idx):
                       data.append([optBegin_idx0,optBegin_idx])
                        
            optBegin_idx0 = optBegin_idx

#去掉重复的内容（数组去重操作）
data1 = [[-1, -1]]
for i in data:
    if not i in data1: 
        data1.append(i) 

wb1 = Workbook()
ws1 = wb1.active

ws1['A1'] = "时间"
ws1['B1'] = "操作"
ws1['C1'] = "端口"
ws1['D1'] = "内容"
        
ws1.column_dimensions['A'].width = 25
ws1.column_dimensions['D'].width = 60

#从list中提取数据        
k = 2;
for i in range(len(data1)):
    print(data1[i])
    if data1[i][0] != -1:
        for j in range(data1[i][0],data1[i][1]):
            key = list[j][0]
            value = list[j][1]
            #print(str(j)+"^"+key)
            #print(str(j)+"^"+value)
            wsWrite(ws1,k,value,j)
            k = k +1
            
file = '重复发送情况.xlsx'
wb1.save(file)
print("异常日志输出至"+file)
        
   
#按照端口+时间进行排序 ok
#如果成功，且无重发 就不要进行考虑   
#如果不成功，需要考虑
#1-2-3-4的  ，且 4 very 为成功的，无需显示 1-2-3-4
#1出现后 4出现  4为成功，不显示

